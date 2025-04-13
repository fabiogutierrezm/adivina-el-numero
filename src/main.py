from utils import dame_aleatorio
import openpyxl
import getpass
import matplotlib.pyplot as plt
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / "data" / "estadisticas_tarea.xlsx"

def modo_solitario(dificultad):
    numerosecreto = dame_aleatorio(1, 1000)
    intentos = dificultad
    while intentos > 0:
        try:
            numerojugado = int(input("Intenta adivinar el número entre el 1 y el 1000: "))
        except ValueError:
            print("Por favor, introduce un número válido.")
            continue

        if numerojugado == numerosecreto:
            print("\n🎉 Enhorabuena, has ganado el juego! 🎉")
            guardar_estadistica("Modo solitario", dificultad, "Ganado")
            return
        elif numerojugado < numerosecreto:
            print("El número a adivinar es mayor.")
        else:
            print("El número a adivinar es menor.")
        intentos -= 1

    print(f"\n😢 Has agotado todos los intentos. El número era {numerosecreto}.")
    guardar_estadistica("Modo solitario", dificultad, "Perdido")

def modo_2jugadores(dificultad):
    try:
        numero_adivinar = int(getpass.getpass("Jugador 1, inserte el número a adivinar entre el 1 y el 1000: "))
    except ValueError:
        print("Número inválido.")
        return

    intentos = dificultad
    while intentos > 0:
        try:
            numerojugado = int(input("Jugador 2, intenta adivinar el número: "))
        except ValueError:
            print("Por favor, introduce un número válido.")
            continue

        if numerojugado == numero_adivinar:
            print("\n🎉 Enhorabuena, has ganado el juego! 🎉")
            guardar_estadistica("Modo 2 jugadores", dificultad, "Ganado")
            return
        elif numerojugado < numero_adivinar:
            print("El número a adivinar es mayor.")
        else:
            print("El número a adivinar es menor.")
        intentos -= 1

    print(f"\n😢 Has agotado los intentos. El número era {numero_adivinar}.")
    guardar_estadistica("Modo 2 jugadores", dificultad, "Perdido")

def guardar_estadistica(mododejuego, dificultad, resultado):
    nombre_jugador = input("Introduce tu nombre: ")
    excel_estadisticas = openpyxl.load_workbook(EXCEL_PATH)
    hoja = excel_estadisticas.active
    dificultad_str = {20: "Fácil", 12: "Media", 5: "Difícil"}.get(dificultad, str(dificultad))
    hoja.append([nombre_jugador, mododejuego, dificultad_str, resultado])
    excel_estadisticas.save(EXCEL_PATH)

def mostrar_estadisticas():
    excel_estadisticas = openpyxl.load_workbook(EXCEL_PATH)
    hoja = excel_estadisticas.active
    ganado, perdido = 0, 0
    print("\n📊 Resultados de las partidas:")
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre, modo_juego, dificultad, resultado = fila
        print(f"Jugador: {nombre}, Modo: {modo_juego}, Dificultad: {dificultad}, Resultado: {resultado}")
        if resultado == "Ganado": ganado += 1
        elif resultado == "Perdido": perdido += 1

    plt.bar(["Ganadas", "Perdidas"], [ganado, perdido], color=['green', 'red'])
    plt.title("Estadísticas de partidas")
    plt.show()

def menu():
    while True:
        print("\n🎮 Bienvenido a Adivina el Número 🎮")
        print("1. Modo solitario")
        print("2. Modo 2 jugadores")
        print("3. Ver estadísticas")
        print("4. Salir")

        try:
            opcion = int(input("Elige una opción: "))
        except ValueError:
            print("Opción inválida. Inténtalo de nuevo.")
            continue

        if opcion in [1, 2]:
            print("\nDificultad: 1.Fácil 2.Media 3.Difícil")
            try:
                nivel = int(input("Selecciona nivel de dificultad: "))
                dificultad = {1: 20, 2: 12, 3: 5}.get(nivel)
                if not dificultad:
                    print("Dificultad no válida.")
                    continue
            except ValueError:
                print("Entrada inválida.")
                continue
            if opcion == 1:
                modo_solitario(dificultad)
            else:
                modo_2jugadores(dificultad)
        elif opcion == 3:
            mostrar_estadisticas()
        elif opcion == 4:
            print("Gracias por jugar. ¡Hasta la próxima!")
            break
        else:
            print("Opción no válida.")

if __name__ == "__main__":
    menu()
